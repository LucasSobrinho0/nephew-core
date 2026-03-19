from pathlib import Path

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import FileResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from companies.repositories import CompanyRepository
from companies.services import CompanyService
from common.documents import normalize_cnpj
from imports.models import ImportJob, ImportJobItem
from imports.repositories import ImportJobItemRepository, ImportJobRepository
from people.repositories import PersonRepository
from people.services import PersonService


class ImportTemplateService:
    TEMPLATE_DIR = Path(settings.BASE_DIR) / 'data' / 'modelos'

    @staticmethod
    def build_headers(entity_type):
        if entity_type == ImportJob.EntityType.PEOPLE:
            return ['nome', 'sobrenome', 'email', 'telefone', 'apollo_id', 'hubspot_id', 'botconversa_id', 'cnpj_empresa']
        if entity_type == ImportJob.EntityType.COMPANIES:
            return ['razao', 'cnpj', 'website', 'email', 'telefone', 'segmento', 'quantidade_funcionarios', 'apollo_id', 'hubspot_id']
        raise ValidationError('Tipo de template invalido.')

    @staticmethod
    def build_template_filename(entity_type):
        return 'person.xlsx' if entity_type == ImportJob.EntityType.PEOPLE else 'companie.xlsx'

    @staticmethod
    def ensure_template_file(entity_type):
        ImportTemplateService.TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
        file_path = ImportTemplateService.TEMPLATE_DIR / ImportTemplateService.build_template_filename(entity_type)
        if file_path.exists():
            return file_path

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'modelo'
        sheet.append(ImportTemplateService.build_headers(entity_type))
        workbook.save(file_path)
        return file_path

    @staticmethod
    def build_download_response(entity_type):
        file_path = ImportTemplateService.ensure_template_file(entity_type)
        return FileResponse(
            file_path.open('rb'),
            as_attachment=True,
            filename=file_path.name,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )


class ImportWorkbookService:
    REQUIRED_HEADERS = {
        ImportJob.EntityType.PEOPLE: {'nome', 'sobrenome'},
        ImportJob.EntityType.COMPANIES: {'razao'},
    }

    @staticmethod
    def normalize_header(value):
        return str(value or '').strip().lower()

    @staticmethod
    def normalize_cell(value):
        if value is None:
            return ''
        return str(value).strip()

    @staticmethod
    def load_rows(file_path, entity_type):
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        headers = [ImportWorkbookService.normalize_header(value) for value in (header_row or [])]
        if not headers or not any(headers):
            raise ValidationError('A planilha XLSX nao possui cabecalho valido.')

        required_headers = ImportWorkbookService.REQUIRED_HEADERS[entity_type]
        missing_headers = sorted(header for header in required_headers if header not in headers)
        if missing_headers:
            raise ValidationError('Cabecalhos obrigatorios ausentes: ' + ', '.join(missing_headers))

        rows = []
        for index, row in enumerate(iterator, start=2):
            payload = {
                header: ImportWorkbookService.normalize_cell(value)
                for header, value in zip(headers, row)
                if header
            }
            if not any(payload.values()):
                continue
            rows.append((index, payload))
        return rows


class ImportJobService:
    STORAGE_DIR = Path(settings.BASE_DIR) / 'data' / 'imports'

    @staticmethod
    def build_storage_path(*, entity_type, public_id):
        ImportJobService.STORAGE_DIR.mkdir(parents=True, exist_ok=True)
        return ImportJobService.STORAGE_DIR / f'{entity_type}-{public_id}.xlsx'

    @staticmethod
    @transaction.atomic
    def create_job(*, user, organization, entity_type, uploaded_file):
        if entity_type not in {ImportJob.EntityType.PEOPLE, ImportJob.EntityType.COMPANIES}:
            raise ValidationError('Tipo de importacao invalido.')

        job = ImportJobRepository.create(
            organization=organization,
            entity_type=entity_type,
            status=ImportJob.Status.PENDING,
            source_filename=uploaded_file.name,
            created_by=user,
            updated_by=user,
        )
        storage_path = ImportJobService.build_storage_path(entity_type=entity_type, public_id=job.public_id)
        with storage_path.open('wb') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)

        rows = ImportWorkbookService.load_rows(storage_path, entity_type)
        if not rows:
            raise ValidationError('A planilha nao possui linhas preenchidas para importar.')

        items = [
            ImportJobItem(
                organization=organization,
                job=job,
                row_number=row_number,
                raw_payload=payload,
            )
            for row_number, payload in rows
        ]
        ImportJobItemRepository.bulk_create(items)

        job.stored_file_path = str(storage_path)
        job.total_rows = len(items)
        job.save(update_fields=['stored_file_path', 'total_rows', 'updated_at'])
        return job

    @staticmethod
    def refresh_counters(*, job, user=None):
        items = ImportJobItemRepository.list_for_job(job)
        job.processed_rows = items.exclude(status=ImportJobItem.Status.PENDING).count()
        job.success_rows = items.filter(status=ImportJobItem.Status.SUCCESS).count()
        job.failed_rows = items.filter(status=ImportJobItem.Status.FAILED).count()
        if user is not None:
            job.updated_by = user

        if job.processed_rows >= job.total_rows and job.total_rows > 0:
            if job.failed_rows and job.success_rows:
                job.status = ImportJob.Status.COMPLETED_WITH_ERRORS
            elif job.failed_rows and not job.success_rows:
                job.status = ImportJob.Status.FAILED
            else:
                job.status = ImportJob.Status.COMPLETED
            job.finished_at = timezone.now()
        job.save(
            update_fields=[
                'processed_rows',
                'success_rows',
                'failed_rows',
                'status',
                'finished_at',
                'updated_by',
                'updated_at',
            ]
        )
        return job


class ImportPeopleService:
    @staticmethod
    def import_payload(*, user, organization, payload):
        first_name = payload.get('nome', '')
        last_name = payload.get('sobrenome', '')
        email = payload.get('email', '')
        phone = payload.get('telefone', '')
        apollo_person_id = payload.get('apollo_id', '')
        hubspot_contact_id = payload.get('hubspot_id', '')
        bot_conversa_id = payload.get('botconversa_id', '')
        company_cnpj = normalize_cnpj(payload.get('cnpj_empresa', ''))

        if not first_name:
            raise ValidationError('A coluna nome e obrigatoria.')
        if not last_name:
            raise ValidationError('A coluna sobrenome e obrigatoria.')

        company = None
        if company_cnpj:
            company = CompanyRepository.get_for_organization_and_cnpj(organization, company_cnpj)
            if company is None:
                raise ValidationError(f'Nenhuma empresa local encontrada para o CNPJ {company_cnpj}.')

        person = PersonService.create_person(
            user=user,
            organization=organization,
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            email=email,
            bot_conversa_id=bot_conversa_id,
            hubspot_contact_id=hubspot_contact_id,
            apollo_person_id=apollo_person_id,
            company=company,
        )
        return person.full_name


class ImportCompaniesService:
    @staticmethod
    def import_payload(*, user, organization, payload):
        name = payload.get('razao', '')
        cnpj = payload.get('cnpj', '')
        website = payload.get('website', '')
        email = payload.get('email', '')
        phone = payload.get('telefone', '')
        segment = payload.get('segmento', '')
        employee_count_raw = payload.get('quantidade_funcionarios', '')
        apollo_company_id = payload.get('apollo_id', '')
        hubspot_company_id = payload.get('hubspot_id', '')

        if not name:
            raise ValidationError('A coluna razao e obrigatoria.')

        employee_count = None
        if employee_count_raw:
            try:
                employee_count = int(employee_count_raw)
            except (TypeError, ValueError) as exc:
                raise ValidationError('Quantidade de funcionarios invalida.') from exc

        company = CompanyService.create_company(
            user=user,
            organization=organization,
            name=name,
            cnpj=cnpj,
            website=website,
            email=email,
            phone=phone,
            segment=segment,
            employee_count=employee_count,
            apollo_company_id=apollo_company_id,
            hubspot_company_id=hubspot_company_id,
        )
        return company.name


class ImportJobWorkerService:
    ENTITY_HANDLERS = {
        ImportJob.EntityType.PEOPLE: ImportPeopleService.import_payload,
        ImportJob.EntityType.COMPANIES: ImportCompaniesService.import_payload,
    }

    @staticmethod
    @transaction.atomic
    def process_job(*, job, batch_size=20):
        if job.status in {ImportJob.Status.COMPLETED, ImportJob.Status.COMPLETED_WITH_ERRORS, ImportJob.Status.FAILED}:
            return job

        if job.status == ImportJob.Status.PENDING:
            job.status = ImportJob.Status.RUNNING
            job.started_at = job.started_at or timezone.now()
            job.save(update_fields=['status', 'started_at', 'updated_at'])

        handler = ImportJobWorkerService.ENTITY_HANDLERS[job.entity_type]
        pending_items = list(ImportJobItemRepository.list_pending_for_job(job, limit=batch_size))

        if not pending_items:
            return ImportJobService.refresh_counters(job=job, user=job.updated_by or job.created_by)

        actor = job.updated_by or job.created_by
        for item in pending_items:
            try:
                result_label = handler(
                    user=actor,
                    organization=job.organization,
                    payload=item.raw_payload,
                )
                item.status = ImportJobItem.Status.SUCCESS
                item.message = f'Importado com sucesso: {result_label}'[:255]
            except ValidationError as exc:
                messages = getattr(exc, 'messages', None) or [str(exc)]
                item.status = ImportJobItem.Status.FAILED
                item.message = (messages[0] or 'Falha na validacao da linha.')[:255]
            item.save(update_fields=['status', 'message', 'updated_at'])

        return ImportJobService.refresh_counters(job=job, user=actor)

    @staticmethod
    def run_cycle(*, limit=20, batch_size=20):
        jobs = list(ImportJobRepository.list_runnable_jobs(limit=limit))
        processed_count = 0
        for job in jobs:
            ImportJobWorkerService.process_job(job=job, batch_size=batch_size)
            processed_count += 1
        return processed_count


class ImportJobPresentationService:
    @staticmethod
    def build_payload(job):
        items = ImportJobItemRepository.list_for_job(job)
        return {
            'status': job.status,
            'progress_percent': job.progress_percent,
            'processed_rows': job.processed_rows,
            'total_rows': job.total_rows,
            'success_rows': job.success_rows,
            'failed_rows': job.failed_rows,
            'is_finished': job.status in {
                ImportJob.Status.COMPLETED,
                ImportJob.Status.COMPLETED_WITH_ERRORS,
                ImportJob.Status.FAILED,
            },
            'items': [
                {
                    'row_number': item.row_number,
                    'status': item.status,
                    'message': item.message,
                }
                for item in items
            ],
        }
